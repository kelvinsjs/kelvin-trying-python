from os import startfile
from openpyxl.styles import PatternFill

from openpyxl import load_workbook

book = load_workbook('1.xlsx')
sheet = book['start']
rows = sheet.rows
headers = [cell.value for cell in next(rows)]

rows = sheet.rows

all_data = []

for row in rows:
    data = {}
    for title, cell in zip(headers, row):
        data[title] = cell.value
    all_data.append(data)

all_data.pop(0)   
sheet = book['finish']

for i in range(73):
    sheet.cell(row=2, column=i+2).value = 365+i*5
    
for i in range(36):
    sheet.cell(row=i+3, column=3).value = 38+i


def printInCell():
    for i in all_data:
        print(f"iteration # {i}")
        sheetColumn = int(float(f'{i["colX"]}')) - 113 + 3
        sheetRow = int((int(float(i["colY"]))- 365)/5 + 3)
        sheet.cell(row=sheetColumn, column=sheetRow).value = i["colZ"]

printInCell()

newData = []
newRows = sheet.rows
newheaders = [cell.value for cell in next(newRows)]

for row in newRows:
    data = {}
    for title, cell in zip(newheaders, row):
        data[title] = cell.value
    newData.append(data)
    
print(newData)
            

print("Done!")
book.save("1.xlsx")
